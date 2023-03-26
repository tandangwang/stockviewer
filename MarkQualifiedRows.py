"""
初次逻辑筛选，今日股价上涨，且在蜡烛图中，今日上涨柱整体包裹住昨日下跌柱。

最右侧增加1列，字段名设置为 “开收差”，其计算方法是 “收开差”=“收盘”- “开盘”。
筛选条件：
1.当前行“开收差”为正值;
2.当前行“开盘”小于上一行“收盘”;
3.当前行“收盘”大于上一行“开盘”;

结果写入新的xlsx文件。
"""

import openpyxl
import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment


def auto_center_and_adjust_width(ws):
    # 居中对齐
    for row in ws.rows:
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # 自动调整列宽
    for column in ws.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width


class MarkQualifiedRows:
    def __init__(self):
        self.output_filename = ''
        self.prompt = ''

    def get_output_filename(self):
        return self.output_filename

    def get_prompt(self):
        return self.prompt

    def run(self, filename):
        # 读取原始Excel文件
        df = pd.read_excel(filename)

        # 新增“开收差”列并计算数值
        df["开收差"] = df["收盘"] - df["开盘"]

        # 根据“开收差”的正负值对整行进行标记
        def highlight(row):
            if row["开收差"] > 0:
                return ["background-color: red" for i in row]
            elif row["开收差"] < 0:
                return ["background-color: green" for i in row]
            else:
                return ["background-color: white" for i in row]

        df_style = df.style.apply(highlight, axis=1)

        # 新增“初次筛选结果”列并计算数值
        df["初次筛选结果"] = "no"
        for i in range(1, len(df)):
            if df.loc[i, "开收差"] > 0 and df.loc[i, "开盘"] < df.loc[i - 1, "收盘"] and df.loc[i, "收盘"] > df.loc[
                i - 1, "开盘"]:
                df.loc[i, "初次筛选结果"] = "yes"

        # 将DataFrame写入Excel文件中
        self.output_filename = f"{filename}_markrows.xlsx"
        with pd.ExcelWriter(self.output_filename, engine='openpyxl') as writer:
            writer.book = writer.book = openpyxl.Workbook()
            df.to_excel(writer, index=False, sheet_name='Sheet1')

            # 获取Sheet1的worksheet对象
            ws = writer.book["Sheet1"]
            auto_center_and_adjust_width(ws)

            # 遍历DataFrame的每一行，根据“开收差”的正负值为单元格添加填充颜色
            for r in range(2, len(df) + 2):
                cell = ws.cell(row=r, column=df.columns.get_loc("开收差") + 1)
                if cell.value > 0:
                    fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
                    for c in range(1, len(df.columns) + 1):
                        ws.cell(row=r, column=c).fill = fill
                elif cell.value < 0:
                    fill = PatternFill(start_color='FF00FF00', end_color='FF00FF00', fill_type='solid')
                    for c in range(1, len(df.columns) + 1):
                        ws.cell(row=r, column=c).fill = fill

            # 遍历DataFrame的每一行，为“初次筛选结果”列添加填充颜色
            for r in range(2, len(df) + 2):
                cell = ws.cell(row=r, column=df.columns.get_loc("初次筛选结果") + 2)
                if cell.value == "yes":
                    fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
                    cell.fill = fill

        self.prompt = '筛选及行标记完成'
