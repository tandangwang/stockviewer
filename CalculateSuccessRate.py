"""
基于v1.0逻辑筛选，在结果为 yes 的日期买入，持有指定天数后，统计投资涨幅达到预期的成功率。
"""

import openpyxl


def letter_index(lt):
    """
    convert letter to column number
    :param lt:
    :return:
    """
    lts = 'ABCDEFGHIGKLMNOPQRSTUVWXYZ'
    idx = lts.find(lt) + 1
    return idx


def field_name_to_column_index(sheet, column_name):
    """
    字段名转为列序号（int）
    :param sheet:
    :param column_name:
    :return:
    """
    # 遍历第1行中的每个单元格
    for cell in sheet[1]:
        # 获取单元格的列号
        column_number = cell.column
        # 获取单元格的值
        value = cell.value
        # 如果值为“天空”，则输出对应的列号
        if value == column_name:
            return column_number


def count_value_in_column(sheet, value_, column_number):
    """
    统计指定列中，值为v的个数
    :param sheet:
    :param value_: 可为任意格式，如字符串或int
    :param column_number: 行号
    :return:
    """
    # 遍历每一行
    count = 0
    for row in sheet.iter_rows():
        # 获取指定列的单元格
        cell = row[column_number - 1]
        # 获取单元格的值
        value = cell.value
        # 如果值为“天空”，则增加计数器
        if value == value_:
            count += 1
    return count


def float_to_percentage(float_value):
    """
    将小数转换为百分比
    :param float_value:
    :return:
    """
    string_percent = f"{round(float_value * 100, 2)}%"
    return string_percent


class CalculateSuccessRate:
    def __init__(self,target,days_hold):
        self.target = 0.03
        self.days_hold = 10
        self.investment_dic = {}
        self.success_rate = 0
        self.prompt = ''

    def get_prompt(self):
        return self.prompt

    def run(self, filename):
        # 读取xlsx文件
        workbook = openpyxl.load_workbook(filename)
        worksheet = workbook['Sheet1']

        column_open_number = field_name_to_column_index(worksheet, '开盘')
        column_high_number = field_name_to_column_index(worksheet, '最高')
        column_date_number = field_name_to_column_index(worksheet, '日期')
        column_diff_number = field_name_to_column_index(worksheet, '初次筛选结果')

        # 遍历每一行
        # for i, row in enumerate(worksheet.iter_rows()):
        # 从第2行开始，因为第1行是字段名
        for i in range(2, worksheet.max_row):
            # 获取当前行的行号
            current_row_number = i
            current_row = worksheet[current_row_number]

            # 只遍历 初次筛选结果 为 yes 的行
            cell_value = worksheet.cell(row=current_row_number, column=column_diff_number).value
            if cell_value != 'yes':
                continue

            # 获取向下第 持有天数 的行号，第1天是买入日
            target_row_number = current_row_number + self.days_hold

            # 买入价为开盘价
            price_buy = worksheet.cell(row=current_row_number, column=column_open_number).value
            # 卖出价为最高价
            price_sell = worksheet.cell(row=target_row_number, column=column_high_number).value

            # 防止持有时间超过数据边界
            if price_sell is None:
                continue

            # 计算价格增长率 real_percentage
            price_increase = price_sell - price_buy
            increase_percentage = price_increase / price_buy

            if increase_percentage >= self.target:
                date = current_row[column_date_number - 1].value
                # if type(value)==str:
                #     print('stop')
                # 注意current_row是tuple，不是openpyxl的内置类型，所以index要减1
                # date = current_row[column_date_number - 1].value.strftime('%Y-%m-%d')
                self.investment_dic[date] = [{'持有天数': self.days_hold}, {'升值比例': float_to_percentage(increase_percentage)}]

        # 计算success_rate
        yes_count = count_value_in_column(worksheet, value_='yes',
                                          column_number=field_name_to_column_index(worksheet,
                                                                                                   column_name='初次筛选结果'))
        investment_num = len(self.investment_dic)
        if investment_num > 0:
            self.success_rate = len(self.investment_dic) / yes_count

        # 输出结果
        print(self.investment_dic)
        print("投资成功率: " + float_to_percentage(self.success_rate))

        self.prompt = "投资成功率: " + float_to_percentage(self.success_rate)
